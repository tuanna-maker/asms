#!/usr/bin/env python3
"""Xuất use case ASMS ra Excel theo mẫu HRM_QuanLyXe."""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
UC_MD = ROOT / "docs/file docs/use-case-asms.md"
OUT_XLSX = ROOT / "docs/file docs/ASMS_152_UseCases.xlsx"

HIDDEN_MODULE_KEYS = {"de-tai", "cong-viec", "dao-tao"}

SUBMODULE_LABELS = {
    "san-pham.tong-quan": "Tổng quan sản phẩm",
    "san-pham.linh-kien": "Linh kiện / BOM",
    "san-pham.thong-so": "Thông số kỹ thuật",
    "san-pham.tai-lieu": "Tài liệu sản phẩm",
    "san-pham.lich-su": "Lịch sử thay đổi",
    "san-pham.dao-tao": "Đào tạo trên SP",
    "vat-tu.kho": "Kho vật tư",
    "vat-tu.dieu-chuyen": "Phiếu điều chuyển",
    "khach-hang.khach-hang": "Hồ sơ khách hàng",
    "khach-hang.lien-he": "Liên hệ",
    "khach-hang.hoat-dong": "Hoạt động CRM",
    "khach-hang.loyalty": "Kỷ niệm / loyalty",
    "bao-cao.khach-hang": "Báo cáo khách hàng",
    "bao-cao.hop-dong": "Báo cáo hợp đồng",
    "bao-cao.dong-sp": "Báo cáo dòng SP",
    "bao-cao.phan-anh": "Báo cáo phản ánh",
    "bao-cao.don-vi": "Báo cáo đơn vị",
    "cai-dat.nguoi-dung": "Người dùng",
    "cai-dat.vai-tro": "Vai trò",
    "cai-dat.phan-quyen": "Phân quyền",
    "cai-dat.thong-bao": "Thông báo cá nhân",
    "cai-dat.he-thong": "Hệ thống",
    "cai-dat.phien": "Phiên đăng nhập",
    "cai-dat.nhat-ky": "Nhật ký audit",
    "cai-dat.thuoc-tinh": "Danh mục thuộc tính",
}


def humanize_submodule(value: str | None) -> str:
    if not value or value in {"—", "-"}:
        return "Chức năng chính"
    clean = value.strip().strip("`")
    if clean in SUBMODULE_LABELS:
        return SUBMODULE_LABELS[clean]
    if "." in clean:
        return clean.split(".", 1)[1].replace("-", " ").title()
    return clean.replace("-", " ").title()


def classify_action(name: str, perm: str = "") -> str:
    n = name.lower()
    if any(k in n for k in ("đăng nhập", "đăng xuất", "làm mới phiên", "thu hồi phiên")):
        return "Xác thực"
    if any(k in n for k in ("xem danh sách", "xem chi tiết", "xem tổng quan", "xem ", "tra cứu")):
        return "Xem / Tra cứu"
    if any(k in n for k in ("lọc", "tìm kiếm")):
        return "Tìm kiếm"
    if any(k in n for k in ("tạo", "nhập", "thêm", "đăng ký")):
        return "Tạo mới"
    if any(k in n for k in ("sửa", "cập nhật", "điền", "gán", "quản lý", "đánh dấu", "cấu hình", "sắp xếp")):
        return "Cập nhật"
    if "xóa" in n or perm == "delete":
        return "Hủy / Xóa"
    if any(k in n for k in ("xử lý quy trình", "duyệt", "phê duyệt", "đóng", "mở lại", "phân công", "phân luồng")):
        return "Phê duyệt / Xử lý"
    if any(k in n for k in ("xuất", "in ", "thống kê", "báo cáo")):
        return "Báo cáo / Xuất"
    if "nhận" in n or "tự động" in n:
        return "Tự động"
    if perm == "read":
        return "Xem / Tra cứu"
    if perm == "create":
        return "Tạo mới"
    if perm in {"update", "CRUD"}:
        return "Cập nhật"
    return "Xem / Tra cứu"


def plain_actor(actor: str) -> str:
    mapping = {
        "Mọi user": "Người dùng",
        "Theo quyền": "Người dùng được phân quyền",
        "Admin": "Quản trị viên",
        "Hệ thống": "Hệ thống",
        "Admin / quản lý": "Quản trị viên hoặc quản lý",
        "Hệ thống + user": "Hệ thống và người dùng",
    }
    actor = (actor or "Người dùng").strip()
    if actor.startswith("`") or ("." in actor and " " not in actor) or actor in {"—", "-"}:
        return "Người dùng được phân quyền"
    return mapping.get(actor, actor)


def expand_name(name: str) -> str:
    text = name.strip()
    replacements = {
        "HĐ": "hợp đồng",
        "BH": "bảo hành",
        "HL": "huấn luyện",
        "SP": "sản phẩm",
        "KH": "khách hàng",
        "VT": "vật tư",
        "PA": "phản ánh",
        "QT": "quy trình",
        "CV": "công việc",
    }
    for old, new in replacements.items():
        text = re.sub(rf"\b{old}\b", new, text)
    return text[0].lower() + text[1:] if text else text


def plain_note(text: str) -> str:
    if not text:
        return ""
    note = text.strip()
    if note.lower().startswith("submodule:"):
        return ""

    direct_notes = {
        "Email/mật khẩu → JWT (`POST /api/v1/auth/login`)": "Dùng email và mật khẩu để vào hệ thống",
        "Thu hồi refresh token": "Kết thúc phiên đăng nhập hiện tại",
        "Refresh access token": "Gia hạn thời gian đăng nhập",
        "Bootstrap hoặc `POST /api/v1/users`": "Tạo tài khoản người dùng mới",
        "Admin: tất cả; khác: chỉ được phân công": "Quản trị viên xem tất cả; người khác chỉ xem phản ánh được giao",
    }
    if note in direct_notes:
        return direct_notes[note]

    replacements = {
        "Admin:": "Quản trị viên:",
        "JWT": "đăng nhập",
        "refresh token": "phiên đăng nhập",
        "access token": "phiên đăng nhập",
        "soft delete": "xóa khỏi danh sách",
        "CRUD": "thêm, sửa, xóa",
        "RBAC": "phân quyền",
        "SLA": "thời hạn xử lý",
        "BOM": "danh sách linh kiện",
        "CRM": "chăm sóc khách hàng",
        "PAKD": "phản ánh khách hàng",
        "PA": "phản ánh",
        "HĐ": "hợp đồng",
        "BH": "bảo hành",
        "HL": "huấn luyện",
        "SP": "sản phẩm",
        "KH": "khách hàng",
        "VT": "vật tư",
        "QT": "quy trình",
    }
    for old, new in replacements.items():
        note = note.replace(old, new)

    note = re.sub(r"`[^`]*`", "", note)
    note = re.sub(r"\b(?:GET|POST|PUT|PATCH|DELETE)\s+/api/v1/[^\s,)]+", "", note, flags=re.I)
    note = re.sub(r"/[\w\-/:]+", "", note)
    note = re.sub(r"\b(?:read|create|update|delete)\b", "", note, flags=re.I)
    note = re.sub(r"[()→]", " ", note)
    note = re.sub(r"\s+", " ", note).strip(" ,;.")
    if len(note) < 8 or note.lower() in {"bootstrap hoặc", "hoặc"}:
        return ""
    return note


def useful_note(note: str, name: str, sentence: str) -> bool:
    if not note:
        return False
    low = note.lower()
    sent = sentence.lower()
    if low in sent or sent in low:
        return False
    if any(word in low for word in sent.split() if len(word) > 4) and len(low) < 40:
        overlap = sum(1 for word in low.split() if word in sent)
        if overlap >= 2:
            return False
    if any(bad in low for bad in ("bootstrap", "api", "patch ", "query ", "ui export", "usage check", "reorder")):
        return False
    return True


def build_description(uc: dict, mod: dict) -> str:
    actor = plain_actor(uc.get("actor") or "Người dùng")
    action = classify_action(uc["name"], uc.get("perm", ""))
    name = expand_name(uc["name"])
    module = mod["title"]
    group = humanize_submodule(uc.get("submodule"))

    if action == "Xác thực" and "đăng nhập" in name:
        sentence = f"{actor} đăng nhập vào hệ thống bằng email và mật khẩu."
    elif action == "Xác thực" and "đăng xuất" in name:
        sentence = f"{actor} đăng xuất khỏi hệ thống."
    elif action == "Xác thực" and "phiên" in name:
        sentence = f"{actor} quản lý các phiên đăng nhập trên thiết bị."
    elif action == "Tự động":
        sentence = f"Hệ thống tự động thực hiện: {name}."
    elif action == "Xem / Tra cứu":
        sentence = f"{actor} xem và tra cứu: {name}."
    elif action == "Tìm kiếm":
        sentence = f"{actor} tìm kiếm và lọc dữ liệu theo nhu cầu: {name}."
    elif action == "Tạo mới":
        sentence = f"{actor} tạo mới: {name}."
    elif action == "Cập nhật":
        sentence = f"{actor} chỉnh sửa hoặc cập nhật: {name}."
    elif action == "Hủy / Xóa":
        sentence = f"{actor} xóa bản ghi: {name}."
    elif action == "Phê duyệt / Xử lý":
        sentence = f"{actor} xử lý nghiệp vụ: {name}."
    elif action == "Báo cáo / Xuất":
        sentence = f"{actor} xem, xuất hoặc in: {name}."
    else:
        sentence = f"{actor} thực hiện: {name}."

    if group != "Chức năng chính":
        sentence = sentence.rstrip(".") + f", tại mục {group}."

    sentence = sentence.rstrip(".") + f". Thuộc phân hệ {module}."

    note = plain_note(uc.get("desc", ""))
    if useful_note(note, uc["name"], sentence):
        sentence += f" {note}."

    return sentence


def parse_uc_file(md_text: str) -> list[dict]:
    section_re = re.compile(r"^## (\d+)\.\s+(.+?)(?:\s+`([^`]+)`)?(?:\s+—\s+(.+))?$", re.M)
    sections = []
    for m in section_re.finditer(md_text):
        title = m.group(2).strip()
        module_key = m.group(3)
        key_in_title = re.search(r"\(`([^`]+)`\)", title)
        if key_in_title:
            module_key = key_in_title.group(1)
            title = re.sub(r"\s*\(`[^`]+`\)", "", title).strip()
        sections.append(
            {
                "index": m.start(),
                "num": m.group(1),
                "title": title,
                "module_key": module_key,
                "note": m.group(4),
            }
        )

    modules: list[dict] = []
    for i, sec in enumerate(sections):
        start = sec["index"]
        end = sections[i + 1]["index"] if i + 1 < len(sections) else len(md_text)
        block = md_text[start:end]
        title_line = block.split("\n", 1)[0]
        if any(x in title_line for x in ("Tổng hợp", "Luồng nghiệp vụ", "Tài liệu liên quan", "Quy ước")):
            continue

        route_match = re.search(r"\*\*Route UI:\*\*\s+(.+)", block)
        api_match = re.search(r"\*\*API:\*\*\s+(.+)", block)

        ucs: list[dict] = []
        for row in re.findall(r"^\| UC-[^\n]+\|", block, re.M):
            if "Mã UC" in row or "---" in row:
                continue
            cols = [c.strip() for c in row.split("|") if c.strip()]
            if len(cols) < 3:
                continue
            code, name = cols[0], cols[1]
            actor = ""
            desc = ""
            perm = ""
            submodule = ""

            if len(cols) == 3:
                third = cols[2]
                if third.startswith("`") or third.startswith("bao-cao.") or third.startswith("cai-dat.") or third.startswith("san-pham.") or third.startswith("vat-tu.") or third.startswith("khach-hang."):
                    submodule = third.strip("`")
                    actor = "Theo quyền"
                else:
                    actor = third
            elif len(cols) >= 4 and cols[2].startswith("`") and "." in cols[2]:
                submodule = cols[2].strip("`")
                perm = cols[3]
                actor = "Theo quyền"
            elif len(cols) >= 4 and cols[2] in {"read", "create", "update", "delete", "CRUD"}:
                perm = cols[2]
                actor = "Theo quyền"
                desc = cols[3] if len(cols) > 3 else ""
            elif len(cols) >= 4 and cols[3] in {"read", "create", "update", "delete", "CRUD"}:
                actor = cols[2] if cols[2] not in {"—", "-"} else "Theo quyền"
                perm = cols[3]
                if cols[2].startswith("`") or "." in cols[2]:
                    submodule = cols[2].strip("`")
                    actor = "Theo quyền"
                desc = cols[4] if len(cols) > 4 else ""
            elif len(cols) >= 4:
                actor = cols[2]
                extra = cols[3]
                if extra.startswith("`") or extra.startswith("bao-cao.") or extra.startswith("cai-dat."):
                    submodule = extra.strip("`")
                else:
                    desc = extra
                if len(cols) > 4:
                    desc = (desc + " " + cols[4]).strip() if desc else cols[4]
            else:
                actor = cols[2]

            ucs.append(
                {
                    "code": code,
                    "name": name,
                    "actor": actor,
                    "desc": desc,
                    "perm": perm,
                    "submodule": submodule,
                }
            )

        if not ucs:
            continue

        modules.append(
            {
                "num": sec["num"],
                "title": sec["title"],
                "module_key": sec["module_key"] or module_key,
                "note": sec["note"],
                "route": route_match.group(1).strip() if route_match else "",
                "api": api_match.group(1).strip() if api_match else "",
                "ucs": ucs,
            }
        )

    return modules


def is_hidden(mod: dict) -> bool:
    note = (mod.get("note") or "").lower()
    key = mod.get("module_key") or ""
    return "ẩn" in note or key in HIDDEN_MODULE_KEYS


def sheet_name(title: str) -> str:
    name = title.replace("/", "-").strip()
    return name[:31]


HEADER_FILL = PatternFill("solid", fgColor="334155")
TITLE_FILL = PatternFill("solid", fgColor="1E293B")
THIN = Side(style="thin", color="CBD5E0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_header_row(ws, row: int) -> None:
    for col in range(1, 9):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def style_title_row(ws, row: int, text: str) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, color="FFFFFF", size=16)
    cell.fill = TITLE_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")


def style_subtitle_row(ws, row: int, text: str) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, size=11)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_group_row(ws, row: int, text: str) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, size=10, color="1E3A5F")
    cell.fill = PatternFill("solid", fgColor="EEF4FF")


def write_data_row(ws, row: int, values: list, bold_code: bool = True) -> None:
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.border = BORDER
        cell.alignment = WRAP if col in (3, 8) else CENTER if col == 1 else Alignment(vertical="top")
        if col == 2 and bold_code:
            cell.font = Font(bold=True, size=10)


def set_column_widths(ws) -> None:
    widths = [6, 12, 42, 28, 28, 14, 16, 48]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def uc_row(stt: int, uc: dict, mod: dict) -> list:
    module_label = f"ASMS — {mod['title']}"
    group = humanize_submodule(uc.get("submodule"))
    action = classify_action(uc["name"], uc.get("perm", ""))
    return [
        stt,
        uc["code"],
        uc["name"],
        module_label,
        group,
        "Cao",
        action,
        build_description(uc, mod),
    ]


def write_summary_sheet(wb, modules: list[dict]) -> None:
    ws = wb.active
    ws.title = "Tổng hợp"
    total = sum(len(m["ucs"]) for m in modules)
    parts = "   |   ".join(f"{len(m['ucs'])} UC {m['title']}" for m in modules)

    style_title_row(ws, 1, f"ASMS — TỔNG HỢP {total} USE CASE")
    style_subtitle_row(ws, 2, parts)
    headers = ["STT", "Mã UC", "Tên chức năng", "Module", "Nhóm màn hình", "Mức độ ưu tiên", "Loại thao tác", "Mô tả"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=3, column=col, value=h)
    style_header_row(ws, 3)

    row = 4
    stt = 1
    for mod in modules:
        for uc in mod["ucs"]:
            write_data_row(ws, row, uc_row(stt, uc, mod))
            row += 1
            stt += 1
    set_column_widths(ws)
    ws.freeze_panes = "A4"


def write_module_sheet(wb, mod: dict) -> None:
    ws = wb.create_sheet(sheet_name(mod["title"]))
    count = len(mod["ucs"])
    style_title_row(ws, 1, f"{mod['title'].upper()} — DANH SÁCH USE CASE")
    style_subtitle_row(ws, 2, f"▶  {mod['title']}  ({count} use case)")
    headers = ["STT", "Mã UC", "Tên chức năng", "Module", "Nhóm màn hình", "Mức độ ưu tiên", "Loại thao tác", "Mô tả"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=3, column=col, value=h)
    style_header_row(ws, 3)

    groups: dict[str, list[dict]] = {}
    for uc in mod["ucs"]:
        group = humanize_submodule(uc.get("submodule"))
        groups.setdefault(group, []).append(uc)

    row = 4
    stt = 1
    for group_name, ucs in groups.items():
        style_group_row(ws, row, f"  ▸  {group_name}")
        row += 1
        for uc in ucs:
            write_data_row(ws, row, uc_row(stt, uc, mod))
            row += 1
            stt += 1
    set_column_widths(ws)
    ws.freeze_panes = "A4"


def main() -> None:
    md_text = UC_MD.read_text(encoding="utf-8")
    all_modules = parse_uc_file(md_text)
    modules = [m for m in all_modules if not is_hidden(m)]
    total = sum(len(m["ucs"]) for m in modules)

    wb = openpyxl.Workbook()
    write_summary_sheet(wb, modules)
    for mod in modules:
        write_module_sheet(wb, mod)

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"Exported {total} use cases -> {OUT_XLSX}")


if __name__ == "__main__":
    main()
